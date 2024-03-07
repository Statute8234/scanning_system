import cv2
from pyzbar.pyzbar import decode
from openpyxl import load_workbook

# file
path = r"C:\Users\danie\Desktop\New Microsoft Excel Worksheet.xlsx"
workbook = load_workbook(path)
worksheet = workbook.active

# input
user = input('what wold you like to do? (find item | scan item)? > ')
if user == 'scan item':
    # cam
    cap = cv2.VideoCapture(0)
    cap.set(3,640)
    cap.set(4,480)
    used_codes = []
    total_inv = []

    # add to used codes list
    for i in range(1,worksheet.max_row + 1):
        used_codes.append(worksheet.cell(i,2).value)

    # add to inv
    list = []
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            list.append(col[i].value)
        total_inv.append(list)
        list = []

    try:
        while True:
            success, frame = cap.read()

            for code in decode(frame):
                if code.data.decode('utf-8') not in used_codes:
                    used_codes.append(code.data.decode('utf-8'))
                    name = input('what is the name of this item? ')
                    code = code.data.decode('utf-8')
                    location = input('were is this located? ')
                    worksheet.append([name,code,location])
                    workbook.save(path)

            cv2.imshow('Testing-code-scan',frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

    except KeyboardInterrupt:
        quit(1)

else:
    while True:
        look = input('what would you like to look for? > ')
        line = ''
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if col[i].value != 'none':
                    line += str(col[i].value) + ' '

            if look not in line:
                line = ''
            else:
                print('found {} at {}'.format(look,line))
